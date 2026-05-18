from __future__ import annotations

TEST_MODE = False

import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Literal, TypedDict

from langgraph.graph import StateGraph

def load_project_documents():
    from pathlib import Path

    docs_path = Path("inputs/docs")
    combined_text = ""

    if not docs_path.exists():
        return ""

    for file in docs_path.glob("*"):
        try:
            combined_text += f"\n\n--- DOCUMENT: {file.name} ---\n"

            # TXT / MD
            if file.suffix.lower() in [".txt", ".md"]:
                combined_text += file.read_text(encoding="utf-8")

            # PDF
            elif file.suffix.lower() == ".pdf":
                try:
                    from pypdf import PdfReader
                    reader = PdfReader(str(file))
                    for page in reader.pages:
                        combined_text += page.extract_text() or ""
                except Exception as e:
                    combined_text += f"[PDF read error: {e}]"

            # Word DOCX
            elif file.suffix.lower() == ".docx":
                try:
                    from docx import Document
                    doc = Document(str(file))
                    for para in doc.paragraphs:
                        combined_text += para.text + "\n"
                except Exception as e:
                    combined_text += f"[DOCX read error: {e}]"

            else:
                combined_text += f"[Unsupported file type: {file.suffix}]"

        except Exception as e:
            combined_text += f"\nError reading {file.name}: {e}\n"

    return combined_text

import os
import streamlit as st
from openai import OpenAI

api_key = None

try:
    api_key = st.secrets["OPENAI_API_KEY"]
except Exception:
    pass

if not api_key:
    api_key = os.environ.get("OPENAI_API_KEY")

client = OpenAI(api_key=api_key)

MODEL_FAST = "gpt-4.1-mini"

MODEL_FAST = "gpt-4.1-mini"
MODEL_DEEP = "gpt-5.4"

Mode = Literal["deterministic", "reasoning", "deep", "regular_plus_deep"]

class State(TypedDict, total=False):
    topic: str
    main_question: str
    hypotheses: str
    key_angles: str
    data_ideas: str
    user_sources: str
    audience: str
    tone_style: str
    success_criteria: str
    output_constraints: str
    report_type: str
    report_template: dict
    mode: Mode

    project_documents: str
    project_instructions: str
    research_approach_instructions: str
    research_coach_principles: str
    research_coach_notes: str
    methodology_question: str
    research_depth: str
    include_data_sources: bool
    include_systems_analysis: bool
    methodology_notes: str

    route: str
    plan: str
    research_a: str
    research_b: str
    research_c: str

    evidence_table: str
    evidence_check: str
    fact_check: str
    systems_multidisciplinary_analysis: str

    evidence_quality_a: str
    evidence_quality_b: str
    evidence_quality_c: str
    evidence_quality_disagreements: str
    evidence_quality_consensus: str

    synthesis: str
    run_summary: str
    saved_file: str


def ensure_folders():
    Path("templates").mkdir(exist_ok=True)
    Path("outputs").mkdir(exist_ok=True)


def choose_model(mode: Mode):
    return MODEL_DEEP if mode in ("deep", "regular_plus_deep") else MODEL_FAST


def call_llm(system_prompt, user_prompt, mode, use_web=True):
    if use_web:
        system_prompt += """

ADDITIONAL RESEARCH INSTRUCTIONS:
- Use uploaded project documents as primary sources when relevant.
- Search for additional evidence beyond the uploaded documents.
- Prioritize:
  1. Official government sources
  2. Peer-reviewed studies and systematic reviews
  3. Academic working papers
  4. Credible nonpartisan organizations
  5. High-quality journalism
  6. Advocacy materials only to characterize stakeholder claims
"""

    kwargs = {
        "model": choose_model(mode),
        "input": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    }

    if use_web:
        kwargs["tools"] = [{"type": "web_search"}]
        kwargs["tool_choice"] = "auto"

    response = client.responses.create(**kwargs)
    return response.output_text.strip()

def safe_input(prompt, default):
    if TEST_MODE:
        print(f"{prompt}: using default -> {default}")
        return default
    val = input(f"{prompt} [{default}]: ").strip()
    return val if val else default


# =========================
# INTAKE
# =========================
def collect_inputs():
    return {
        "topic": safe_input("Topic", "Colorado fentanyl penalties ballot measure"),
        "main_question": safe_input("Main question", "What does evidence suggest?"),
        "hypotheses": safe_input("Hypotheses", "Penalties may deter but also create tradeoffs"),
        "key_angles": safe_input("Angles", "deterrence, treatment, costs"),
        "data_ideas": safe_input("Data ideas", "overdose trends, incarceration"),
        "user_sources": safe_input("Preferred sources", "government, academic"),
        "audience": safe_input("Audience", "policy staff"),
        "tone_style": safe_input("Tone", "neutral, clear"),
        "success_criteria": safe_input("Success criteria", "balanced, evidence-based"),
        "output_constraints": safe_input("Constraints", "no overclaiming"),
        "report_type": "policy_memo",
        "mode": safe_input("Mode", "reasoning"),
        "revision_count": 0,
    }


# =========================
# NODES
# =========================
def triage(state):
    return {"route": "standard"}


def planner(state):
    prompt = f"Create a research plan for: {state['topic']}"
    return {"plan": call_llm("planner", prompt, state["mode"])}


def baseline(state):
    return {
        "research_a": call_llm(
            "baseline",
            f"Analyze: {state['topic']}",
            state["mode"],
            use_web=True,
        )
    }


def alternative(state):
    return {
        "research_b": call_llm(
            "alternative",
            f"Alternative view: {state['topic']}",
            state["mode"],
            use_web=True,
        )
    }


def skeptical(state):
    return {
        "research_c": call_llm(
            "skeptical",
            f"Critique: {state['topic']}",
            state["mode"],
            use_web=True,
        )
    }

def evidence_table(state):
    prompt = f"""
Extract structured evidence from the research outputs below.

RESEARCH OUTPUTS:
{state['research_a']}

{state['research_b']}

{state['research_c']}

Return ONLY a valid JSON list.

Each item must include:
- claim
- evidence
- source_title
- source_url
- confidence
- limitations
- evidence_status

STRICT REQUIREMENTS:
- evidence_status must be one of: strong, limited, missing
- If no real source exists, write source_needed for source_title and source_url
- Do NOT invent sources
- If evidence is weak or indirect, mark confidence as low or medium
- Always include limitations
- Prefer government, academic, legislative, and high-quality research sources
- Explicitly flag evidence gaps rather than smoothing them over
- You MUST include weak or missing evidence entries where applicable

Example format:
[
  {{
    "claim": "Specific claim",
    "evidence": "Brief evidence summary",
    "source_title": "Source title or source_needed",
    "source_url": "URL or source_needed",
    "confidence": "high / medium / low",
    "limitations": "Important caveats or source gaps",
    "evidence_status": "strong / limited / missing"
  }}
]
"""

    return {
        "evidence_table": call_llm(
            "evidence",
            prompt,
            state["mode"],
            use_web=True,
        )
    }

STRICT_EVIDENCE_RUBRIC = """
Evidence Level Definitions - STRICT:

Level 5 - Proven / strong causal evidence:
- Multiple high-quality RCTs, systematic reviews, meta-analyses, or strong quasi-experimental studies.
- Clear causal identification.
- Replicated or strongly supported across settings.

Level 4 - Evidence-informed / moderate causal evidence:
- Strong quasi-experimental or well-designed observational studies.
- Methods may include difference-in-differences, instrumental variables, regression discontinuity, propensity score matching, strong comparison groups, or similar approaches.
- Supports cautious causal claims, but with limitations.

Level 3 - Theory-informed, rigorous:
- Strong descriptive or correlational evidence.
- Good data and reasonable methods, such as regression with controls, panel data, fixed effects, or structured comparative analysis.
- Suggestive but not strong causal evidence.

Level 2 - Theory-informed, limited:
- Descriptive data, trend analysis, cross-sectional analysis, simple before/after comparisons, single-source analysis, or basic modeling.
- Useful for context, but weak for causal claims.

Level 1 - Weak / speculative:
- Expert opinion, advocacy claims, stakeholder statements, case examples, anecdotes, or claims with limited empirical support.
- Can describe perspectives, but should not be treated as evidence of effectiveness.

No Level - Descriptive only:
- Pure factual claims, such as:
  - a bill was introduced
  - a measure contains a provision
  - an agency reported a statistic
  - a fiscal note estimated a cost
- These claims may have high factual certainty, but they are not evidence of effectiveness or impact.

Important distinction:
- Evidence Level answers: What kind of evidence supports effectiveness, impact, or causal inference?
- Certainty answers: How confident should we be that the claim is accurate and relevant?

Rules:
- Do not assign Level 4 or Level 5 to news articles merely because they are credible.
- Do not assign Level 4 or Level 5 to draft policy language unless the claim is purely descriptive, and then use No Level - Descriptive only.
- Descriptive legal, fiscal, or statistical claims can have High certainty but should usually be No Level - Descriptive only unless evaluating impact.
- Advocacy claims should usually be Level 1 or No Level unless supported by independent empirical evidence.
- Expert opinion without cited empirical support should usually be Level 1 or No Level, with Low or Very Low certainty for impact claims.
- Causal claims require causal evidence.
"""

def evidence_quality_agent_a(state):
    prompt = f"""
You are Evidence Quality Reviewer A.

Use the strict evidence rubric below.

{STRICT_EVIDENCE_RUBRIC}

PROJECT INSTRUCTIONS:
{state.get('project_instructions', '')}

PROJECT DOCUMENTS:
{state.get('project_documents', '')}

EVIDENCE TABLE:
{state['evidence_table']}

Assess each item and return a table with:
- claim
- source
- evidence level
- certainty
- audit rationale
- limitations
- appropriate use in final report

Important:
- Provide an audit rationale, not hidden chain-of-thought.
- Be clear about why each rating was assigned.
- Separate factual certainty from causal/effectiveness evidence.
"""

    return {
        "evidence_quality_a": call_llm(
            "evidence_quality_reviewer_a",
            prompt,
            state["mode"],
            use_web=False,
        )
    }

def evidence_quality_agent_b(state):
    prompt = f"""
You are Evidence Quality Reviewer B.

Independently assess the evidence using the strict rubric below.
Be more conservative than Reviewer A.

{STRICT_EVIDENCE_RUBRIC}

PROJECT INSTRUCTIONS:
{state.get('project_instructions', '')}

PROJECT DOCUMENTS:
{state.get('project_documents', '')}

EVIDENCE TABLE:
{state['evidence_table']}

Downgrade evidence if it is:
- only news reporting
- only expert opinion
- advocacy-based
- indirect
- descriptive rather than causal
- missing methods
- not specific to the research question
- based on projections rather than observed outcomes

Return a table with:
- claim
- source
- evidence level
- certainty
- audit rationale
- limitations
- appropriate use in final report

Important:
- Provide an audit rationale, not hidden chain-of-thought.
- Be conservative where evidence is indirect, descriptive, or non-causal.
"""

    return {
        "evidence_quality_b": call_llm(
            "evidence_quality_reviewer_b",
            prompt,
            state["mode"],
            use_web=False,
        )
    }

def evidence_quality_agent_c(state):
    prompt = f"""
You are Evidence Quality Reviewer C.

Assess the evidence from a practical public-facing policy-analysis perspective.
Use the strict rubric below.

{STRICT_EVIDENCE_RUBRIC}

PROJECT INSTRUCTIONS:
{state.get('project_instructions', '')}

PROJECT DOCUMENTS:
{state.get('project_documents', '')}

EVIDENCE TABLE:
{state['evidence_table']}

Focus on whether each claim is strong enough to support:
- a factual background statement
- a tentative/exploratory insight
- an evidence-supported finding
- a final conclusion

Return a table with:
- claim
- source
- evidence level
- certainty
- audit rationale
- limitations
- appropriate use in final report

Important:
- Provide an audit rationale, not hidden chain-of-thought.
- Explain whether each claim supports causal language, association only, background context, or should be treated as uncertain.
"""

    return {
        "evidence_quality_c": call_llm(
            "evidence_quality_reviewer_c",
            prompt,
            state["mode"],
            use_web=False,
        )
    }

def evidence_quality_disagreement_review(state):
    prompt = f"""
Compare the three independent evidence quality reviews.

REVIEWER A:
{state['evidence_quality_a']}

REVIEWER B:
{state['evidence_quality_b']}

REVIEWER C:
{state['evidence_quality_c']}

Identify:
1. Where reviewers agree
2. Where reviewers disagree
3. Which disagreements matter most
4. Which evidence ratings should be reviewed by a human
5. Whether any reviewer appears too generous or too strict

Return a clear disagreement memo for human review.
"""

    return {
        "evidence_quality_disagreements": call_llm(
            "evidence_quality_disagreement_review",
            prompt,
            state["mode"],
            use_web=False,
        )
    }


def evidence_quality_consensus(state):
    prompt = f"""
Create a consensus evidence-quality assessment.

Use:
- the original evidence table
- Reviewer A assessment
- Reviewer B assessment
- Reviewer C assessment
- disagreement memo

EVIDENCE TABLE:
{state['evidence_table']}

REVIEWER A:
{state['evidence_quality_a']}

REVIEWER B:
{state['evidence_quality_b']}

REVIEWER C:
{state['evidence_quality_c']}

DISAGREEMENT MEMO:
{state['evidence_quality_disagreements']}

Return a final consensus table with:
- claim
- source
- consensus evidence level
- consensus certainty
- reason for rating
- key limitations
- human review flag: yes/no
- appropriate use in final report

Rules:
- Be conservative.
- If reviewers disagree meaningfully, choose the lower-confidence rating unless there is a clear reason not to.
- Do not present Level 1, No Level, Low, or Very Low certainty evidence as strong evidence.
"""

    return {
        "evidence_quality_consensus": call_llm(
            "evidence_quality_consensus",
            prompt,
            state["mode"],
            use_web=False,
        )
    }


def evidence_check(state):
    return {
        "evidence_check": call_llm(
            "check",
            state["evidence_table"],
            state["mode"],
        )
    }

def source_priority_fact_check(state):
    prompt = f"""
You are a general source-priority and fact-check reviewer.

EVIDENCE TABLE:
{state['evidence_table']}

BASELINE:
{state['research_a']}

ALTERNATIVE:
{state['research_b']}

SKEPTICAL:
{state['research_c']}

SOURCE PRIORITY:
1. Primary sources (laws, official documents, datasets)
2. Government / official reports
3. Academic literature
4. Research institutions
5. News
6. Advocacy / opinion

TASK:
- Identify claims needing primary source confirmation
- Flag weak sourcing
- Suggest corrections
- Identify claims to treat as uncertain

Return:
- High-priority fact checks
- Weak source issues
- Suggested corrections
- Claims that are solid
"""

    return {
        "fact_check": call_llm("fact_check", prompt, state["mode"], use_web=True)
    }


def systems_and_multidisciplinary_analysis(state):
    prompt = f"""
You are a multidisciplinary systems thinker.

Use mental models only where they materially improve the analysis. Consider lenses such as incentives, tradeoffs, base rates, feedback loops, opportunity costs, implementation capacity, behavioral responses, institutional incentives, and second- or third-order effects—but do not force every lens.

TOPIC:
{state['topic']}

EVIDENCE:
{state['evidence_table']}

FACT CHECK:
{state.get('fact_check', '')}

TASK:
Develop:
1. Exploratory insights — creative, hypothesis-generating, clearly marked as tentative
2. Evidence-supported insights — grounded in the evidence table and fact-check
3. Second- and third-order effects — only where plausible and relevant
4. Unintended consequences and implementation risks
5. Non-obvious implications for decision-makers

Rules:
- Do not force a checklist of every mental model.
- Prioritize insights that could change how a decision-maker understands the issue.
- Clearly distinguish tentative ideas from evidence-supported findings.
- Be practical, concise, and evidence-aware.
"""

    return {
        "systems_multidisciplinary_analysis": call_llm(
            "systems_multidisciplinary_analysis",
            prompt,
            state["mode"]
        )
    }

def synthesis(state):
    prompt = f"""
Write the final report.

Use the evidence, fact-check, evidence-quality consensus, and systems analysis below.

EVIDENCE TABLE:
{state['evidence_table']}

EVIDENCE QUALITY CONSENSUS:
{state.get('evidence_quality_consensus', '')}

EVIDENCE QUALITY DISAGREEMENTS FOR HUMAN REVIEW:
{state.get('evidence_quality_disagreements', '')}

SOURCE-PRIORITY FACT CHECK:
{state.get('fact_check', '')}

SYSTEMS + MULTIDISCIPLINARY ANALYSIS:
{state.get('systems_multidisciplinary_analysis', '')}

RESEARCH COACH NOTES:
{state.get('research_coach_notes', '')}

METHODOLOGY NOTES:
{state.get('methodology_notes', '')}
f
REQUIREMENTS:
- Follow the report template and do not force every possible section.
- Clearly distinguish three categories:
  1. Exploratory insights: creative or hypothesis-generating; label as tentative.
  2. Evidence-supported findings: sourced, rated, and grounded in the evidence-quality consensus.
  3. Final conclusions: no stronger than the evidence permits.
- Use the consensus evidence-quality ratings when weighing claims.
- Automatically downgrade weak evidence in conclusions.
- Do not present Level 1, No Level, Low, or Very Low certainty evidence as strong findings.
- Clearly label weak evidence, opinion-based evidence, advocacy-based evidence, and indirect evidence.
- Include systems or multidisciplinary insights only where they materially improve the report.
- Include second- and third-order effects only where relevant and plausible.
- Include key evidence gaps.
- Use the Research Coach Notes where they materially improve the report.
- Use the Methodology Notes where they help explain how the system approaches the analysis.
- Avoid overclaiming.
- Use practical, clear, evidence-based language.

Suggested structure, but adapt to the report type:
1. Bottom line
2. Evidence-supported findings
3. Evidence quality and source support
4. Exploratory or tentative insights
5. Key evidence gaps
6. Relevant systems effects / unintended consequences
7. Tradeoffs
8. Limitations
9. Sources
10. Human-review items
"""

    return {
        "synthesis": call_llm("synthesis", prompt, state["mode"])
    }

def save_output(state):
    import csv
    import json
    from pathlib import Path
    from datetime import datetime

    Path("outputs").mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    report_file = Path("outputs") / f"report_{timestamp}.md"
    csv_file = Path("outputs") / f"evidence_export_{timestamp}.csv"
    xlsx_file = Path("outputs") / f"evidence_export_{timestamp}.xlsx"

    report = f"""
# Final Report

{state['synthesis']}

---

# Appendix A: Evidence Table

{state.get('evidence_table', '')}

---

# Appendix B: Evidence Quality Consensus

{state.get('evidence_quality_consensus', '')}

---

# Appendix C: Evidence Quality Disagreements for Human Review

{state.get('evidence_quality_disagreements', '')}

---

# Appendix D: Source-Priority Fact Check

{state.get('fact_check', '')}

---

# Appendix E: Systems + Multidisciplinary Analysis

{state.get('systems_multidisciplinary_analysis', '')}

---

# Appendix F: Evidence Quality Reviewer A

{state.get('evidence_quality_a', '')}

---

# Appendix G: Evidence Quality Reviewer B

{state.get('evidence_quality_b', '')}

---

# Appendix H: Evidence Quality Reviewer C

{state.get('evidence_quality_c', '')}
"""

    report_file.write_text(report, encoding="utf-8")

    # Try to export the evidence table to CSV / Excel-ready format.
    # This works best when evidence_table is valid JSON.
    evidence_rows = []

    try:
        raw = state.get("evidence_table", "").strip()

        # Try direct JSON first.
        try:
            evidence_rows = json.loads(raw)
        except Exception:
            # Try extracting JSON array from model output.
            start = raw.find("[")
            end = raw.rfind("]")
            if start != -1 and end != -1 and end > start:
                evidence_rows = json.loads(raw[start:end + 1])

        if not isinstance(evidence_rows, list):
            evidence_rows = []

    except Exception:
        evidence_rows = []

    fields = [
        "claim",
        "evidence",
        "source_title",
        "source_url",
        "confidence",
        "limitations",
        "evidence_status",
    ]

    # Always write CSV so Excel can open it.
    with csv_file.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()

        if evidence_rows:
            for row in evidence_rows:
                writer.writerow({field: row.get(field, "") for field in fields})
        else:
            writer.writerow({
                "claim": "Evidence table was not valid JSON",
                "evidence": "See Markdown report appendix for evidence table",
                "source_title": "",
                "source_url": "",
                "confidence": "",
                "limitations": "CSV export requires evidence_table to be valid JSON",
                "evidence_status": "",
            })

    # Try to also write a true .xlsx file if openpyxl is installed.
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Evidence"

        ws.append(fields)

        if evidence_rows:
            for row in evidence_rows:
                ws.append([row.get(field, "") for field in fields])
        else:
            ws.append([
                "Evidence table was not valid JSON",
                "See Markdown report appendix for evidence table",
                "",
                "",
                "",
                "XLSX export requires evidence_table to be valid JSON",
                "",
            ])

        # Add a second sheet for evidence quality consensus text.
        ws2 = wb.create_sheet("Quality Consensus")
        ws2["A1"] = state.get("evidence_quality_consensus", "")

        # Add a third sheet for human review disagreements.
        ws3 = wb.create_sheet("Human Review")
        ws3["A1"] = state.get("evidence_quality_disagreements", "")

        wb.save(xlsx_file)

    except Exception:
        # If openpyxl is not installed, CSV is still created.
        xlsx_file = "Not created. To enable .xlsx export, run: pip install openpyxl"

    print("\nSaved report:", report_file)
    print("Saved evidence CSV:", csv_file)
    print("Saved Excel file:", xlsx_file)

    return {
        "saved_file": str(report_file),
        "saved_evidence_csv": str(csv_file),
        "saved_excel_file": str(xlsx_file),
    }

def research_coach(state):
    prompt = f"""
You are acting as a senior research advisor for the Institute of Evidence-Based Policymaking.

Your role is to help the researcher think more deeply and systematically about the issue.

TOPIC:
{state.get('topic', '')}

MAIN QUESTION:
{state.get('main_question', '')}

PROJECT INSTRUCTIONS:
{state.get('project_instructions', '')}

CUSTOM RESEARCH COACH PRINCIPLES:
{state.get('research_coach_principles', '')}

PLANNING NOTES:
{state.get('plan', '')}

EVIDENCE QUALITY CONSENSUS:
{state.get('evidence_quality_consensus', '')}

SYSTEMS AND MULTIDISCIPLINARY ANALYSIS:
{state.get('systems_multidisciplinary_analysis', '')}

DEFAULT PRINCIPLES TO APPLY IF NOT OVERRIDDEN:
- Seek disconfirming evidence.
- Distinguish clearly among facts, stakeholder claims, evidence-supported findings, and conclusions.
- Focus on causal mechanisms.
- Identify key assumptions that drive the conclusion.
- Consider incentives, tradeoffs, and opportunity costs.
- Explore plausible second- and third-order effects.
- Highlight implementation challenges and capacity constraints.
- Identify what evidence would most change the conclusion.
- Surface key uncertainties and evidence gaps.
- Consider the strongest arguments on all sides.

Prepare a section titled "Research Coach Notes" with the following sections:

1. Key Questions to Consider
2. Assumptions to Test
3. Strongest Arguments on Each Side
4. Potential Blind Spots
5. Important Evidence Gaps
6. Suggested Data Sources and How to Obtain Them
7. Additional Research Questions
8. Issues Requiring Human Judgment
9. Recommended Next Steps

Distinguish clearly between:
- Exploratory insights
- Evidence-supported findings
- Open questions

Use a thoughtful, practical tone intended to help a researcher refine the project and deepen understanding.
"""

    return {
        "research_coach_notes": call_llm(
            "research coach",
            prompt,
            state["mode"],
            use_web=False,
        )
    }

def methodology_explainer(state):
    methodology_question = state.get("methodology_question", "").strip()

    # If no methodology question was provided, skip.
    if not methodology_question:
        return {"methodology_notes": ""}

    prompt = f"""
You are explaining how the Institute's multi-agent research system works.

TOPIC:
{state.get('topic', '')}

METHODOLOGY QUESTION:
{methodology_question}

SYSTEM DESIGN FEATURES:
- Multiple research agents investigate the issue from different perspectives.
- A structured evidence table is created.
- Three independent evidence-quality reviewers assess the strength and certainty of evidence.
- A disagreement-review step identifies and summarizes differences among reviewers.
- A consensus step reconciles disagreements.
- A source-priority fact check emphasizes highly credible sources.
- Systems and multidisciplinary analysis identifies tradeoffs and unintended consequences.
- A Research Coach generates key questions, assumptions, blind spots, and suggested data sources.
- Final synthesis distinguishes:
  1. Exploratory insights
  2. Evidence-supported findings
  3. Final conclusions

EVIDENCE FRAMEWORK:
- Stronger causal and more directly relevant evidence receives greater weight.
- Advocacy-based, speculative, indirect, and weak evidence is explicitly downgraded.
- Conclusions are no stronger than the underlying evidence permits.

TASK:
Answer the methodology question clearly and specifically.
Explain how the current system addresses the issue.
Note any limitations or areas where human judgment remains important.

Use a concise, transparent, and practical tone.
"""

    return {
        "methodology_notes": call_llm(
            "methodology explainer",
            prompt,
            state["mode"],
            use_web=False,
        )
    }

# =========================
# BUILD GRAPH
# =========================

builder = StateGraph(State)

builder.add_node("triage", triage)
builder.add_node("planner", planner)
builder.add_node("baseline", baseline)
builder.add_node("alternative", alternative)
builder.add_node("skeptical", skeptical)
builder.add_node("evidence_table", evidence_table)

builder.add_node("evidence_quality_a", evidence_quality_agent_a)
builder.add_node("evidence_quality_b", evidence_quality_agent_b)
builder.add_node("evidence_quality_c", evidence_quality_agent_c)
builder.add_node("evidence_quality_disagreement_review", evidence_quality_disagreement_review)
builder.add_node("evidence_quality_consensus", evidence_quality_consensus)

builder.add_node("evidence_check", evidence_check)
builder.add_node("source_priority_fact_check", source_priority_fact_check)
builder.add_node("systems_and_multidisciplinary_analysis", systems_and_multidisciplinary_analysis)
builder.add_node("research_coach", research_coach)
builder.add_node("methodology_explainer", methodology_explainer)
builder.add_node("synthesis", synthesis)
builder.add_node("save_output", save_output)

builder.set_entry_point("triage")

builder.add_edge("triage", "planner")

builder.add_edge("planner", "baseline")
builder.add_edge("planner", "alternative")
builder.add_edge("planner", "skeptical")

builder.add_edge(["baseline", "alternative", "skeptical"], "evidence_table")

builder.add_edge("evidence_table", "evidence_quality_a")
builder.add_edge("evidence_table", "evidence_quality_b")
builder.add_edge("evidence_table", "evidence_quality_c")

builder.add_edge(
    ["evidence_quality_a", "evidence_quality_b", "evidence_quality_c"],
    "evidence_quality_disagreement_review",
)

builder.add_edge("evidence_quality_disagreement_review", "evidence_quality_consensus")
builder.add_edge("evidence_quality_consensus", "evidence_check")
builder.add_edge("evidence_check", "source_priority_fact_check")
builder.add_edge("source_priority_fact_check", "systems_and_multidisciplinary_analysis")
builder.add_edge("systems_and_multidisciplinary_analysis", "research_coach")
builder.add_edge("research_coach", "methodology_explainer")
builder.add_edge("methodology_explainer", "synthesis")
builder.add_edge("synthesis", "save_output")

builder.set_finish_point("save_output")

graph = builder.compile()

# ▶️ Run the workflow
def run_research(state):
    ensure_folders()
    return graph.invoke(state)
if __name__ == "__main__":
    ensure_folders()
    print("Starting workflow...\n")

    state = {
        "topic": "Colorado 2026 Fentanyl Penalties Ballot Measure (Initiative 85)",

      "main_question": """
Help me think through Colorado Initiative 85.

Explain:
- what the measure would do,
- the problem it is trying to address,
- the strongest arguments on both sides,
- what the best available evidence says,
- important evidence gaps and uncertainties,
- key tradeoffs and unintended consequences,
- and questions researchers and voters should consider.

Ask thoughtful questions that would help refine the research and deepen understanding.
""",

        "mode": "reasoning",

        "project_documents": load_project_documents(),

        "project_instructions": """
Use interactive learning mode.

Act as an evidence-based research partner rather than primarily as a report writer.

Your objectives are to:
- Explain the issue clearly in plain language.
- Identify the key policy, legal, economic, and implementation questions.
- Highlight the strongest arguments and evidence on all sides.
- Suggest useful data sources, research studies, and literature to review.
- Identify important evidence gaps and uncertainties.
- Explore tradeoffs, incentives, and possible second- and third-order effects.
- Ask thoughtful follow-up questions that help refine the research.

Distinguish clearly among:
- descriptive facts,
- stakeholder claims,
- evidence-supported findings,
- exploratory insights,
- final conclusions.

Use the Institute's evidence framework:
- Level 5: Proven / strong causal evidence
- Level 4: Evidence-informed / moderate causal evidence
- Level 3: Theory-informed, rigorous
- Level 2: Theory-informed, limited
- Level 1: Weak / speculative
- No Level: Descriptive only

Evidence certainty:
- High
- Moderate
- Low
- Very Low

Use uploaded documents and external research where helpful.

Do not prioritize producing a polished final report.
The primary goal is to help the researcher think, learn, and refine the questions that should be investigated.
""",
    }

    output = graph.invoke(state)

    print("\nFINAL REPORT:\n")
    print(output.get("synthesis", "No synthesis returned."))

    print("\nSaved files:")
    print("Report:", output.get("saved_file", "Not saved"))
    print("CSV:", output.get("saved_evidence_csv", "Not saved"))
    print("Excel:", output.get("saved_excel_file", "Not saved"))